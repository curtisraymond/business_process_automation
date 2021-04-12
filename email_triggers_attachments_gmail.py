# ----------------------- CREATOR INFO -----------------------
#
# NAME:
# Curtis Raymond
# https://www.linkedin.com/in/curtis-raymond/
#
# ------------------------------------------------------------

import email
import imaplib
import os
import random
import string
import pyautogui
from openpyxl import load_workbook
import traceback
from io import StringIO
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from os.path import basename

# Initial prompt message when this automated process begins
config_file = pyautogui.prompt(text="Enter the PATH of your Excel configuration file", title='email_triggers_attachments_gmail', default='')

# Import the Excel configuration file
wb = load_workbook(config_file)  # Workbook
ws = wb.get_sheet_by_name('config')  # Worksheet

# Your gmail email
gmail_email = ws.cell(row=4, column=5).value

# Your gmail password
password = ws.cell(row=5, column=5).value

# Should email attachment(s) have unique file name(s)? (Yes/No)
unique_file_name = ws.cell(row=6, column=5).value

# Sender's email
from_email = ws.cell(row=7, column=5).value

# Subject of the Sender's email
if ws.cell(row=8, column=5).value == None:
    email_subject = str("')")
else:
    email_subject = str(" SUBJECT ") + str("'") + str(ws.cell(row=8, column=5).value) + str("')")

# Secondary email to send process error issues to
if ws.cell(row=11, column=5).value == None:
    cc_email = ""
else:
    cc_email = ws.cell(row=11, column=5).value

# Where you will save email attachments
attachment_dir = ws.cell(row=9, column=5).value

# Should email(s) detected by this automated process be starred within your gmail?
star_email = ws.cell(row=10, column=5).value

# Connect to the Gmail server
mail = imaplib.IMAP4_SSL("imap.gmail.com")
smtp_ssl_host = 'smtp.gmail.com'
smtp_ssl_port = 465

# Your Gmail credentials
mail.login(gmail_email, password)

# Critical error email template if and when the automated process crashes
critical_error_template = """
Processing has failed. Traceback and error are shown below:

%s
"""


# Automated process when the Gmail IMAP server detects a new email in your inbox
def main_process():
    mail.select("INBOX")
    n = 0
    (retcode, messages) = mail.search(None, "(UNSEEN FROM " + "'" + from_email + email_subject)
    if retcode == 'OK':
        for num in messages[0].split():
            n = n + 1
            print(n)
            typ, data = mail.fetch(num, '(RFC822)')
            for response_part in data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    print(msg['From'])
                    print(msg['Subject'])
                    # Allows you to download email attachments
                    for part in msg.walk():
                        if part.get_content_maintype() == 'multipart':
                            continue
                        if part.get('Content-Disposition') is None:
                            continue
                        file_name = part.get_filename()
                        if bool(file_name):
                            if unique_file_name == "No":
                                file_path = os.path.join(attachment_dir, file_name)
                            elif unique_file_name == "Yes":
                                file_type = file_name[file_name.rfind('.'):]
                                file_name = ''.join(random.choices(string.ascii_uppercase + string.digits, k=30))
                                file_path = os.path.join(attachment_dir, file_name + file_type)
                            with open(file_path, 'wb') as f:
                                f.write(part.get_payload(decode=True))
                    if data == 'eject':
                        ctypes.windll.WINMM.mciSendStringW(u"set cdaudio door open", None, 0, None)
                    if star_email == "Yes":
                        typ, data = mail.store(num, '+FLAGS', '\\Flagged')


# Email trigger if and when the automated process fails/crashes
def process_error(message, take_screenshot):
    mail = MIMEMultipart()
    mail['to'] = gmail_email
    mail['cc'] = cc_email
    mail['subject'] = 'Processing Error!'
    to_address = [gmail_email, cc_email]
    mail.attach(MIMEText(message, 'plain'))
    part = MIMEBase('application', "octet-stream")
    if take_screenshot == True:
        # Takes a screenshot of your computer screen and then adds it to the "error_process" email as an attachment
        my_screenshot = pyautogui.screenshot()
        my_screenshot_path = os.path.join(attachment_dir, 'screenshot.png')
        my_screenshot.save(my_screenshot_path)
        part.set_payload(open(my_screenshot_path, "rb").read())
        encoders.encode_base64(part)
        part['Content-Disposition'] = 'attachment; filename="%s"' % basename(my_screenshot_path)
        mail.attach(part)
        os.remove(my_screenshot_path)
    # Interact with Google's servers to send the email
    server = smtplib.SMTP_SSL(smtp_ssl_host, smtp_ssl_port)
    server.login(gmail_email, password)
    server.sendmail(gmail_email, to_address, mail.as_string())
    server.quit()
    os._exit(1)


# Reconnect to the Gmail server due to this process error: "imaplib.IMAP4.abort: command: SELECT => socket error: EOF"
def login():
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(gmail_email, password)
    return mail


while True:
    try:
        main_process()
    except Exception as e:
        if isinstance(e, mail.abort):
            try:
                mail = login()
                continue
            except:
                exc_string = StringIO()
                traceback.print_exc(file=exc_string)
                message = critical_error_template % (exc_string.getvalue())
                process_error(message, take_screenshot=False)
        else:
            exc_string = StringIO()
            traceback.print_exc(file=exc_string)
            message = critical_error_template % (exc_string.getvalue())
            process_error(message, take_screenshot=False)